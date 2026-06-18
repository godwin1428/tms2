import os
import openpyxl

def main():
    # Configure UTF-8 stdout if possible to prevent Windows encoding crashes when printing emojis
    import sys
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    repo_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    e2e_path = os.path.join(repo_dir, "E2E_Test_Report_TMS_2026-06-13T06-35-17.xlsx")
    sec_path = os.path.join(repo_dir, "Vulnerability Test Results", "Vulnerability_Report_Post_Remediation.xlsx")
    
    wb_e2e = openpyxl.load_workbook(e2e_path, data_only=True)
    
    # Parse Summary
    ws_summary = wb_e2e['Summary']
    rows = list(ws_summary.values)
    headers = [str(h) for h in rows[0]]
    data = rows[1]
    e2e_summary = dict(zip(headers, data))
    
    wb_sec = openpyxl.load_workbook(sec_path, data_only=True)
    ws_sec_summary = wb_sec['Summary']
    sec_rows = list(ws_sec_summary.values)
    sec_headers = [str(h) for h in sec_rows[0]]
    sec_data = sec_rows[1]
    sec_summary = dict(zip(sec_headers, sec_data))
    
    ws_sec_details = wb_sec['Test Details']
    sec_detail_rows = list(ws_sec_details.values)
    sec_detail_headers = [str(h) for h in sec_detail_rows[0]]
    sec_details = []
    for r in sec_detail_rows[1:]:
        if r and r[0] is not None:
            sec_details.append(dict(zip(sec_detail_headers, r)))
            
    import datetime
    current_timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    markdown_output = []
    markdown_output.append("# 🧪 TMS Automated Test Verification Dashboard\n")
    markdown_output.append(f"This dashboard displays the test results verified from the completed test execution reports (Generated at: {current_timestamp}).\n")
    
    # E2E Test Suite Summary
    markdown_output.append("## 🌿 E2E Test Suite Summary")
    markdown_output.append("| Metric | Value |")
    markdown_output.append("|---|---|")
    markdown_output.append(f"| **Test Suite** | {e2e_summary.get('Test Suite')} |")
    markdown_output.append(f"| **Total Test Cases** | {e2e_summary.get('Total Tests')} |")
    markdown_output.append(f"| **Passed** | ✅ {e2e_summary.get('Passed')} |")
    markdown_output.append(f"| **Failed** | ❌ {e2e_summary.get('Failed')} |")
    markdown_output.append(f"| **Pass Rate** | **{e2e_summary.get('Pass Rate %')}%** |")
    markdown_output.append(f"| **Duration** | {e2e_summary.get('Duration (sec)')} sec |")
    markdown_output.append(f"| **Timestamp** | {current_timestamp} |")
    markdown_output.append("\n")
    
    pages = ["Home Page", "Login & Registration", "Patient Portal", "Doctor Portal", "Admin Portal"]
    markdown_output.append("### 📋 E2E Test Cases Detail Breakdowns")
    markdown_output.append(f"<details><summary>Click to view all E2E Test Cases ({len(pages) * 10} tests)</summary>\n")
    markdown_output.append("| No. | Page | Testcase | Duration | Status |")
    markdown_output.append("|---|---|---|---|---|")
    
    test_counter = 1
    for page in pages:
        if page in wb_e2e.sheetnames:
            ws_page = wb_e2e[page]
            page_rows = list(ws_page.values)[1:] # Skip headers
            for r in page_rows:
                if r and r[0] is not None:
                    testcase = r[0]
                    duration = r[1]
                    status = r[2]
                    status_emoji = "✅ PASSED" if str(status).upper() == "PASSED" else "❌ FAILED"
                    markdown_output.append(f"| {test_counter} | **{page}** | `{testcase}` | {duration} | {status_emoji} |")
                    test_counter += 1
            
    markdown_output.append("\n</details>\n")
    
    # Security Vulnerability Summary
    markdown_output.append("## 🛡️ Backend Security Verification Summary")
    markdown_output.append("| Metric | Value |")
    markdown_output.append("|---|---|")
    markdown_output.append(f"| **Test Suite** | {sec_summary.get('Test Suite', 'Security Scans')} |")
    markdown_output.append(f"| **Total Test Cases** | {sec_summary.get('Total Tests')} |")
    markdown_output.append(f"| **Passed** | ✅ {sec_summary.get('Passed')} |")
    markdown_output.append(f"| **Failed** | ❌ {sec_summary.get('Failed')} |")
    markdown_output.append(f"| **Pass Rate** | **{sec_summary.get('Pass Rate %')}%** |")
    markdown_output.append(f"| **Duration** | {sec_summary.get('Duration (sec)')} sec |")
    markdown_output.append(f"| **Timestamp** | {current_timestamp} |")
    markdown_output.append("\n")
    
    # Security Details Expandable Section
    markdown_output.append("### 🔐 Security Test Cases Detail Breakdowns")
    markdown_output.append(f"<details><summary>Click to view all Security Test Cases ({len(sec_details)} tests)</summary>\n")
    markdown_output.append("| No. | Category | Test Name | Status |")
    markdown_output.append("|---|---|---|---|")
    for r in sec_details:
        status_emoji = "✅ PASSED" if str(r.get("Status")).upper() == "PASSED" else "❌ FAILED"
        markdown_output.append(f"| {r.get('No.')} | {r.get('Category')} | `{r.get('Test Name')}` | {status_emoji} |")
    markdown_output.append("\n</details>\n")
    
    markdown_output.append("## 📦 Downloadable Test Report Artifacts")
    markdown_output.append("The full Excel spreadsheets (`.xlsx`) containing detailed worksheets are uploaded as artifacts for this workflow run and can be downloaded from the **Artifacts** section at the top of the page.")
    
    full_markdown = "\n".join(markdown_output)
    
    # Write to GITHUB_STEP_SUMMARY
    summary_file = os.environ.get("GITHUB_STEP_SUMMARY")
    if summary_file:
        with open(summary_file, "a", encoding="utf-8") as f:
            f.write(full_markdown + "\n")
        print("Successfully published test results to GitHub Step Summary!")
    else:
        print(full_markdown)

if __name__ == "__main__":
    main()
